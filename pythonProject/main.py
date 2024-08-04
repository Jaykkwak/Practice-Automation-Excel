import tkinter as tk
import openpyxl
from tkinter import filedialog
from copy import copy

from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.workbook import Workbook

global file_path
global sheet
global column_offset
global row_offset
top_left_cell = 'A1'
dest_wb = Workbook()
dest_ws = dest_wb.active


# Select the Excel file
def select_file():
    global file_path
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename()

    print(file_path)


# Open the Excel file
def open_file():
    global file_path
    global sheet
    # Open the Excel file
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    # Get the sheet you want to work with
    sheet = workbook['DUT']


def get_values():
    global sheet
    # Now you can access individual cells or ranges of cells
    # For example, to read the value of cell A1:
    dut_value = sheet['Y162'].value
    ref_value = sheet['AB162'].value

    # check the value
    if ref_value - dut_value > 5:
        print("FAIL")
        device_info_copy_cells()
        issue_copy_cells()
    else:
        print("PASS")


def device_info_copy_cells():
    global sheet
    global top_left_cell
    global column_offset
    global row_offset
    global dest_ws
    dest_cell = None

    # Define the range to copy
    copy_range = sheet['D16:M23']

    top_left_cell = 'A1'

    # Calculate the column offset directly
    column_offset = column_index_from_string(top_left_cell[0])
    print(column_offset)
    # Calculate the row offset
    row_offset = int(top_left_cell[1:])
    print(row_offset)

    # Iterate over the source range
    for i, row in enumerate(copy_range):
        for j, cell in enumerate(row):
            # Calculate the destination cell coordinates
            dest_cell = dest_ws.cell(row=row_offset + i + 1, column=column_offset + j + 1)
            # Copy value and style from the source cell to the destination cell
            dest_cell.value = cell.value
            if cell.has_style:
                dest_cell.font = copy(cell.font)
                dest_cell.border = copy(cell.border)
                dest_cell.fill = copy(cell.fill)
                dest_cell.number_format = copy(cell.number_format)
                dest_cell.protection = copy(cell.protection)
                dest_cell.alignment = copy(cell.alignment)

        top_left_cell = (dest_cell.row, column_offset + 1)

        # Calculate the cell immediately below the last cell in the destination range
    if top_left_cell:
        top_left_cell = dest_ws.cell(row=top_left_cell[0] + 1, column=top_left_cell[1]).coordinate
        print(top_left_cell)

        dest_wb.save('destination.xlsx')


def issue_copy_cells():
    global sheet
    global top_left_cell
    global column_offset
    global row_offset
    global dest_ws

    # Define the range to copy
    copy_range = sheet['D160:AE162']

    # Calculate the column offset directly
    column_offset = column_index_from_string(top_left_cell[0]) - 2

    # Calculate the row offset
    row_offset = int(top_left_cell[1:])

    # Iterate over the source range
    for i, row in enumerate(copy_range):
        for j, cell in enumerate(row):
            # Calculate the destination cell coordinates
            dest_cell = dest_ws.cell(row=row_offset + i + 1, column=column_offset + j + 1)
            # Copy value and style from the source cell to the destination cell
            dest_cell.value = cell.value
            if cell.has_style:
                dest_cell.font = copy(cell.font)
                dest_cell.border = copy(cell.border)
                dest_cell.fill = copy(cell.fill)
                dest_cell.number_format = copy(cell.number_format)
                dest_cell.protection = copy(cell.protection)
                dest_cell.alignment = copy(cell.alignment)

            top_left_cell = (dest_cell.row, column_offset + 1)

        # Calculate the cell immediately below the last cell in the destination range
    if top_left_cell:
        top_left_cell = dest_ws.cell(row=top_left_cell[0] + 1, column=top_left_cell[1]).coordinate
        print(top_left_cell)

    dest_wb.save('destination.xlsx')


def main():
    select_file()
    open_file()
    get_values()


if __name__ == "__main__":
    main()
